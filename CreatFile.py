"""
Create all excel or other files
"""
import openpyxl, time, glovar as glv
import pandas as pd

gs = glv.global_str()
gss = glv.global_status_str()

def create_report_excel(report_name):
    time_str = time.strftime("%Y%m%d%H%M%S", time.localtime())
    report_file_name = report_name + time_str + '.xlsx'
    save_path = glv.output_file_path + '\\' + report_file_name
    wb = openpyxl.Workbook()
    return wb, save_path


"""
This excel is used to record the Volume Production
* One sheet records one test item
* Signal_Name   DUT_NUM    Test_value
  Signal1       DUT0_num      x
                DUT1_num      x
                ...    
                DUTn_num      x
  ......      
  Signaln       DUT0_num      x
                DUT1_num      x
                ...    
                DUTn_num      x       
  <Sheet_TestItem>
"""
def CreateExcel_VP_log():
    report_name = 'VP_'
    Rcord_excel_pd = glv.marked_df
    shift_count = glv.log_row
    test_name_count = len(glv.tree_checked)
    wb, save_path = create_report_excel(report_name)
    Rcord_excel_pd.to_csv(glv.final_path)
    enter = False
    for index, row in Rcord_excel_pd.iterrows():
        sheet_num = test_name_count
        # if index == 0
        if Rcord_excel_pd.at[index, str(gs.CheckStatus)] == str(gss.Checked):
            if index == 0:  # first line do not need to judge
                # ws = wb.create_sheet(str(Rcord_excel_pd.at[index, str(gs.TestName)]), sheet_num)
                sh_name = wb.sheetnames  # 获取所有sheet
                ws = wb[sh_name[0]]
                ws.title = str(Rcord_excel_pd.at[index, str(gs.TestName)])  # 修改第一个sheet名为dddd
                sheet_num -= 1
                row_num = 0
                ws['A1'] = 'Signal Name'
                ws['B1'] = 'DUT Num'
                ws['C1'] = 'Test Value'
                ws['D1'] = 'Judge Result'
                ws.title = Rcord_excel_pd.at[index, str(gs.TestName)]
                enter = True
            elif Rcord_excel_pd.at[index, str(gs.TestName)] != Rcord_excel_pd.at[index - 1, str(gs.TestName)]:
                ws = wb.create_sheet(str(Rcord_excel_pd.at[index, str(gs.TestName)]), sheet_num)
                sheet_num += 1
                row_num = 0
                ws['A1'] = 'Signal Name'
                ws['B1'] = 'DUT Num'
                ws['C1'] = 'Test Value'
                ws['D1'] = 'Judge Result'
                # ws.column_dimensions['A'].width = 10
                ws.title = Rcord_excel_pd.at[index, str(gs.TestName)]
                enter = True
            executed = False
            if enter:
                for dut in range(glv.dut_count):  # others DUT's log
                    if not executed:
                        ws.cell(row=2 + glv.dut_count*row_num, column=1, value=Rcord_excel_pd.at[
                            index, str(gs.Signal)])  # write signal name in column 1, row++
                        # ws.merge_cells(start_column=1, end_column=1, start_row=2+glv.dut_count*row_num, end_row=2+glv.dut_count*(row_num+1))
                        executed = True
                    ws.cell(row=2 + dut + glv.dut_count*row_num, column=2, value=dut)  # write DUT Number in column 1, row++
                    ws.cell(row=2 + dut + glv.dut_count*row_num, column=3, value=Rcord_excel_pd.at[
                        index + dut * shift_count, str(gs.Measure)])  # write Measure value in column 1, row++
                    ws.cell(row=2 + dut + glv.dut_count*row_num, column=4, value=Rcord_excel_pd.at[
                        index + dut * shift_count, str(gs.Result)])  # write Result in column 1, row++
                row_num += 1
        if index == glv.log_row-1:  # subtract the -EOL- line
            break
    wb.save(save_path)


# create file
def file_create(f_path, f_name, f_type):
    full_path = f_path + f_name + f_type
    fp = open(full_path, 'w')
    return fp


if __name__ == '__main__':
    create_report_excel('ttt')
